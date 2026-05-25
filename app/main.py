from datetime import datetime, time
from io import BytesIO

from fastapi import FastAPI, Request, Depends, Form, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import or_, text, func
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .database import Base, engine, get_db, SessionLocal
from .models import User, Ticket, Comment, Direction
from .auth import hash_password, authenticate_user, get_user_by_username

STATUS_CREATED = "Заявка создана"
STATUS_IN_PROGRESS = "В работе"
STATUS_DONE = "Выполнена"
TICKET_STATUSES = [STATUS_CREATED, STATUS_IN_PROGRESS, STATUS_DONE]

ROLE_USER = "user"
ROLE_ADMIN = "admin"
ROLE_NAMES = {
    ROLE_USER: "Пользователь",
    ROLE_ADMIN: "Администратор",
    "moderator": "Администратор",
}

OLD_STATUS_MAP = {
    "created": STATUS_CREATED,
    "in_progress": STATUS_IN_PROGRESS,
    "done": STATUS_DONE,
    "completed": STATUS_DONE,
}

app = FastAPI(title="Система обработки заявок")
app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")
templates.env.globals["role_name"] = lambda role: ROLE_NAMES.get(role, role)

Base.metadata.create_all(bind=engine)


def ensure_schema():
    """Мини-миграция для уже созданной SQLite-базы без Alembic."""
    with engine.connect() as conn:
        user_cols = [row[1] for row in conn.execute(text("PRAGMA table_info(users)"))]
        ticket_cols = [row[1] for row in conn.execute(text("PRAGMA table_info(tickets)"))]
        if "full_name" not in user_cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN full_name VARCHAR(250) DEFAULT ''"))
        if "created_at" not in user_cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN created_at DATETIME"))
        if "direction_id" not in ticket_cols:
            conn.execute(text("ALTER TABLE tickets ADD COLUMN direction_id INTEGER"))
        conn.commit()


def normalize_existing_data():
    db = SessionLocal()
    try:
        tickets = db.query(Ticket).filter(Ticket.status.in_(list(OLD_STATUS_MAP.keys()))).all()
        for ticket in tickets:
            ticket.status = OLD_STATUS_MAP.get(ticket.status, STATUS_CREATED)

        users = db.query(User).all()
        for user in users:
            if not user.full_name:
                user.full_name = user.username
            if user.role == "moderator":
                user.role = ROLE_ADMIN
            if not user.created_at:
                user.created_at = datetime.utcnow()
        db.commit()
    finally:
        db.close()


def create_default_admin():
    db = SessionLocal()
    try:
        admin = get_user_by_username(db, "admin")
        if not admin:
            db.add(
                User(
                    full_name="Администратор системы",
                    username="admin",
                    password_hash=hash_password("admin123"),
                    role=ROLE_ADMIN,
                )
            )
            db.commit()
    finally:
        db.close()


ensure_schema()
normalize_existing_data()
create_default_admin()


def is_admin(user: User) -> bool:
    return user.role == ROLE_ADMIN or user.role == "moderator"


def get_current_user(request: Request, db: Session = Depends(get_db)):
    user_id = request.cookies.get("user_id")
    if not user_id:
        return None
    try:
        return db.query(User).filter(User.id == int(user_id)).first()
    except ValueError:
        return None


def require_user(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Необходимо войти в систему")
    return user


def require_admin(user: User = Depends(require_user)):
    if not is_admin(user):
        raise HTTPException(status_code=403, detail="Доступ только для администратора")
    return user


def get_ticket_or_404(db: Session, ticket_id: int):
    ticket = (
        db.query(Ticket)
        .options(
            joinedload(Ticket.author),
            joinedload(Ticket.direction),
            joinedload(Ticket.comments).joinedload(Comment.author),
        )
        .filter(Ticket.id == ticket_id)
        .first()
    )
    if not ticket:
        raise HTTPException(status_code=404, detail="Заявка не найдена")
    return ticket


def check_ticket_access(ticket: Ticket, user: User):
    if not is_admin(user) and ticket.author_id != user.id:
        raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")


@app.get("/", response_class=HTMLResponse)
def index(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if user:
        return RedirectResponse("/tickets", status_code=302)
    return templates.TemplateResponse(request=request, name="index.html", context={"user": user})


@app.get("/register", response_class=HTMLResponse)
def register_page(request: Request):
    return templates.TemplateResponse(request=request, name="register.html", context={"user": None, "error": None})


@app.post("/register")
def register(
    request: Request,
    full_name: str = Form(...),
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    full_name = full_name.strip()
    username = username.strip()
    if not full_name:
        return templates.TemplateResponse(
            request=request,
            name="register.html",
            context={"user": None, "error": "Фамилия, имя и отчество обязательны для заполнения"},
        )
    if get_user_by_username(db, username):
        return templates.TemplateResponse(
            request=request,
            name="register.html",
            context={"user": None, "error": "Пользователь уже существует"},
        )
    user = User(full_name=full_name, username=username, password_hash=hash_password(password), role=ROLE_USER)
    db.add(user)
    db.commit()
    return RedirectResponse("/login", status_code=302)


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse(request=request, name="login.html", context={"user": None, "error": None})


@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    user = authenticate_user(db, username, password)
    if not user:
        return templates.TemplateResponse(
            request=request,
            name="login.html",
            context={"user": None, "error": "Неверный логин или пароль"},
        )
    response = RedirectResponse("/tickets", status_code=302)
    response.set_cookie("user_id", str(user.id), httponly=True)
    return response


@app.get("/logout")
def logout():
    response = RedirectResponse("/", status_code=302)
    response.delete_cookie("user_id")
    return response


@app.get("/tickets", response_class=HTMLResponse)
def tickets(
    request: Request,
    search: str = Query(""),
    status_filter: str = Query(""),
    date_from: str = Query(""),
    date_to: str = Query(""),
    author_filter: str = Query(""),
    direction_filter: str = Query(""),
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    query = (
        db.query(Ticket)
        .join(User, Ticket.author_id == User.id)
        .outerjoin(Direction, Ticket.direction_id == Direction.id)
        .options(joinedload(Ticket.author), joinedload(Ticket.direction))
    )

    if not is_admin(user):
        query = query.filter(Ticket.author_id == user.id)

    if status_filter:
        query = query.filter(Ticket.status == status_filter)

    if direction_filter:
        try:
            query = query.filter(Ticket.direction_id == int(direction_filter))
        except ValueError:
            pass

    if search:
        pattern = f"%{search.strip()}%"
        query = query.filter(
            or_(
                Ticket.title.ilike(pattern),
                Ticket.description.ilike(pattern),
                User.username.ilike(pattern),
                User.full_name.ilike(pattern),
                Direction.name.ilike(pattern),
            )
        )

    if date_from:
        try:
            start = datetime.combine(datetime.strptime(date_from, "%Y-%m-%d").date(), time.min)
            query = query.filter(Ticket.created_at >= start)
        except ValueError:
            pass

    if date_to:
        try:
            end = datetime.combine(datetime.strptime(date_to, "%Y-%m-%d").date(), time.max)
            query = query.filter(Ticket.created_at <= end)
        except ValueError:
            pass

    if author_filter and is_admin(user):
        pattern = f"%{author_filter.strip()}%"
        query = query.filter(or_(User.username.ilike(pattern), User.full_name.ilike(pattern)))

    items = query.order_by(Ticket.created_at.desc()).all()
    authors = db.query(User).order_by(User.full_name).all() if is_admin(user) else []
    directions = db.query(Direction).order_by(Direction.name).all()

    return templates.TemplateResponse(
        request=request,
        name="tickets.html",
        context={
            "user": user,
            "tickets": items,
            "statuses": TICKET_STATUSES,
            "authors": authors,
            "directions": directions,
            "filters": {
                "search": search,
                "status_filter": status_filter,
                "date_from": date_from,
                "date_to": date_to,
                "author_filter": author_filter,
                "direction_filter": direction_filter,
            },
        },
    )


@app.get("/tickets/new", response_class=HTMLResponse)
def new_ticket_page(request: Request, db: Session = Depends(get_db), user: User = Depends(require_user)):
    directions = db.query(Direction).order_by(Direction.name).all()
    return templates.TemplateResponse(
        request=request,
        name="ticket_form.html",
        context={"user": user, "ticket": None, "statuses": TICKET_STATUSES, "directions": directions, "error": None},
    )


@app.post("/tickets/new")
def create_ticket(
    request: Request,
    title: str = Form(...),
    description: str = Form(...),
    direction_id: int = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    direction = db.query(Direction).filter(Direction.id == direction_id).first()
    if not direction:
        directions = db.query(Direction).order_by(Direction.name).all()
        return templates.TemplateResponse(
            request=request,
            name="ticket_form.html",
            context={
                "user": user,
                "ticket": None,
                "statuses": TICKET_STATUSES,
                "directions": directions,
                "error": "Выберите направление заявки. Направления добавляет администратор.",
            },
        )
    ticket = Ticket(
        title=title,
        description=description,
        direction_id=direction.id,
        author_id=user.id,
        status=STATUS_CREATED,
    )
    db.add(ticket)
    db.commit()
    return RedirectResponse("/tickets", status_code=302)


@app.get("/tickets/{ticket_id}", response_class=HTMLResponse)
def ticket_detail(ticket_id: int, request: Request, db: Session = Depends(get_db), user: User = Depends(require_user)):
    ticket = get_ticket_or_404(db, ticket_id)
    check_ticket_access(ticket, user)
    return templates.TemplateResponse(
        request=request,
        name="ticket_detail.html",
        context={"user": user, "ticket": ticket, "statuses": TICKET_STATUSES},
    )


@app.post("/tickets/{ticket_id}/comment")
def add_comment(
    ticket_id: int,
    comment_text: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    ticket = get_ticket_or_404(db, ticket_id)
    check_ticket_access(ticket, user)
    text_value = comment_text.strip()
    if text_value:
        db.add(Comment(text=text_value, ticket_id=ticket.id, author_id=user.id))
        db.commit()
    return RedirectResponse(f"/tickets/{ticket_id}", status_code=302)


@app.post("/tickets/{ticket_id}/status")
def change_ticket_status(
    ticket_id: int,
    status_value: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_user),
):
    if not is_admin(user):
        raise HTTPException(status_code=403, detail="Статус может менять только администратор")
    if status_value not in TICKET_STATUSES:
        raise HTTPException(status_code=400, detail="Некорректный статус заявки")
    ticket = get_ticket_or_404(db, ticket_id)
    ticket.status = status_value
    db.commit()
    return RedirectResponse(f"/tickets/{ticket_id}", status_code=302)


@app.get("/tickets/{ticket_id}/edit", response_class=HTMLResponse)
def edit_ticket_page(ticket_id: int):
    return RedirectResponse(f"/tickets/{ticket_id}", status_code=302)


@app.post("/tickets/{ticket_id}/edit")
def edit_ticket(ticket_id: int):
    return RedirectResponse(f"/tickets/{ticket_id}", status_code=302)


@app.post("/tickets/{ticket_id}/delete")
def delete_ticket(ticket_id: int, db: Session = Depends(get_db), user: User = Depends(require_user)):
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Заявка не найдена")
    if not is_admin(user) and ticket.author_id != user.id:
        raise HTTPException(status_code=403, detail="Нет доступа")
    db.delete(ticket)
    db.commit()
    return RedirectResponse("/tickets", status_code=302)


@app.get("/users")
def users_list(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):

    if user.role != "admin":
        return templates.TemplateResponse(
            request=request,
            name="403.html",
            context={
                "user": user
            },
            status_code=403
        )

    users = db.query(User).all()

    return templates.TemplateResponse(
        request=request,
        name="users.html",
        context={
            "user": user,
            "users": users
        }
    )


@app.get("/users/{user_id}", response_class=HTMLResponse)
def user_profile(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_user)
):

    if user.role != "admin" and user.id != user_id:
        return templates.TemplateResponse(
            request=request,
            name="403.html",
            context={
                "user": user
            },
            status_code=403
        )

    profile = db.query(User).filter(
        User.id == user_id
    ).first()

    if not profile:
        raise HTTPException(
            status_code=404,
            detail="Пользователь не найден"
        )

    tickets_count = db.query(Ticket).filter(
        Ticket.author_id == profile.id
    ).count()

    return templates.TemplateResponse(
        request=request,
        name="user_profile.html",
        context={
            "user": user,
            "profile": profile,
            "tickets_count": tickets_count
        },
    )


@app.get("/directions", response_class=HTMLResponse)
def directions_page(request: Request, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    directions = db.query(Direction).order_by(Direction.name).all()
    return templates.TemplateResponse(
        request=request,
        name="directions.html",
        context={"user": user, "directions": directions, "error": None},
    )


@app.post("/directions/add")
def add_direction(
    request: Request,
    name: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    name = name.strip()
    if not name:
        directions = db.query(Direction).order_by(Direction.name).all()
        return templates.TemplateResponse(
            request=request,
            name="directions.html",
            context={"user": user, "directions": directions, "error": "Введите наименование направления"},
        )
    exists = db.query(Direction).filter(Direction.name == name).first()
    if exists:
        directions = db.query(Direction).order_by(Direction.name).all()
        return templates.TemplateResponse(
            request=request,
            name="directions.html",
            context={"user": user, "directions": directions, "error": "Такое направление уже существует"},
        )
    db.add(Direction(name=name))
    db.commit()
    return RedirectResponse("/directions", status_code=302)


@app.post("/directions/{direction_id}/edit")
def edit_direction(
    direction_id: int,
    name: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    direction = db.query(Direction).filter(Direction.id == direction_id).first()
    if not direction:
        raise HTTPException(status_code=404, detail="Направление не найдено")
    name = name.strip()
    if name:
        direction.name = name
        db.commit()
    return RedirectResponse("/directions", status_code=302)


@app.post("/directions/{direction_id}/delete")
def delete_direction(direction_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    direction = db.query(Direction).filter(Direction.id == direction_id).first()
    if not direction:
        raise HTTPException(status_code=404, detail="Направление не найдено")
    has_tickets = db.query(Ticket).filter(Ticket.direction_id == direction.id).first()
    if has_tickets:
        raise HTTPException(status_code=400, detail="Нельзя удалить направление, которое используется в заявках")
    db.delete(direction)
    db.commit()
    return RedirectResponse("/directions", status_code=302)


@app.get("/admin", response_class=HTMLResponse)
def admin_panel(request: Request, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    users = db.query(User).order_by(User.id).all()
    return templates.TemplateResponse(request=request, name="admin.html", context={"user": user, "users": users})


@app.post("/admin/users/{user_id}/role")
def change_role(user_id: int, role: str = Form(...), db: Session = Depends(get_db), user: User = Depends(require_admin)):
    if role not in [ROLE_USER, ROLE_ADMIN]:
        raise HTTPException(status_code=400, detail="Некорректная роль")
    target = db.query(User).filter(User.id == user_id).first()
    if target and target.username != "admin":
        target.role = role
        db.commit()
    return RedirectResponse("/admin", status_code=302)

REPORT_COLUMNS = {
    "id": "ID",
    "title": "Тема",
    "description": "Описание",
    "status": "Статус",
    "direction": "Направление",
    "author": "Автор",
    "created_at": "Дата создания",
    "comments_count": "Кол-во комментариев",
}

REPORT_PRESETS = {
    "summary": "Сводка по заявкам",
    "detailed": "Детальный список заявок",
    "by_status": "Группировка по статусам",
    "by_direction": "Группировка по направлениям",
    "by_author": "Группировка по пользователям",
    "custom": "Пользовательский отчёт",
}


def build_report_data(
    db: Session,
    report_type: str,
    date_from: str = "",
    date_to: str = "",
    status_filter: str = "",
    direction_filter: str = "",
    author_filter: str = "",
    selected_columns: list[str] | None = None,
):
    query = (
        db.query(Ticket)
        .join(User, Ticket.author_id == User.id)
        .outerjoin(Direction, Ticket.direction_id == Direction.id)
        .options(
            joinedload(Ticket.author),
            joinedload(Ticket.direction),
            joinedload(Ticket.comments),
        )
    )

    if status_filter:
        query = query.filter(Ticket.status == status_filter)

    if direction_filter:
        try:
            query = query.filter(Ticket.direction_id == int(direction_filter))
        except ValueError:
            pass

    if author_filter:
        pattern = f"%{author_filter.strip()}%"
        query = query.filter(
            or_(
                User.username.ilike(pattern),
                User.full_name.ilike(pattern),
            )
        )

    if date_from:
        try:
            start = datetime.combine(datetime.strptime(date_from, "%Y-%m-%d").date(), time.min)
            query = query.filter(Ticket.created_at >= start)
        except ValueError:
            pass

    if date_to:
        try:
            end = datetime.combine(datetime.strptime(date_to, "%Y-%m-%d").date(), time.max)
            query = query.filter(Ticket.created_at <= end)
        except ValueError:
            pass

    tickets = query.order_by(Ticket.created_at.desc()).all()

    total = len(tickets)

    status_stats = {}
    direction_stats = {}
    author_stats = {}

    for ticket in tickets:
        status_stats[ticket.status] = status_stats.get(ticket.status, 0) + 1

        direction_name = ticket.direction.name if ticket.direction else "Не указано"
        direction_stats[direction_name] = direction_stats.get(direction_name, 0) + 1

        author_name = ticket.author.full_name if ticket.author else "Не указан"
        author_stats[author_name] = author_stats.get(author_name, 0) + 1

    if not selected_columns:
        selected_columns = ["id", "title", "status", "direction", "author", "created_at"]

    rows = []

    if report_type in ["detailed", "custom"]:
        for ticket in tickets:
            row = {}
            for column in selected_columns:
                if column == "id":
                    row[column] = ticket.id
                elif column == "title":
                    row[column] = ticket.title
                elif column == "description":
                    row[column] = ticket.description
                elif column == "status":
                    row[column] = ticket.status
                elif column == "direction":
                    row[column] = ticket.direction.name if ticket.direction else "Не указано"
                elif column == "author":
                    row[column] = ticket.author.full_name if ticket.author else "Не указан"
                elif column == "created_at":
                    row[column] = ticket.created_at.strftime("%d.%m.%Y %H:%M") if ticket.created_at else ""
                elif column == "comments_count":
                    row[column] = len(ticket.comments)
            rows.append(row)

    return {
        "total": total,
        "tickets": tickets,
        "status_stats": status_stats,
        "direction_stats": direction_stats,
        "author_stats": author_stats,
        "rows": rows,
        "selected_columns": selected_columns,
    }


@app.get("/reports", response_class=HTMLResponse)
def reports_page(
    request: Request,
    report_type: str = Query("summary"),
    date_from: str = Query(""),
    date_to: str = Query(""),
    status_filter: str = Query(""),
    direction_filter: str = Query(""),
    author_filter: str = Query(""),
    columns: list[str] = Query(default=[]),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    directions = db.query(Direction).order_by(Direction.name).all()

    report_data = build_report_data(
        db=db,
        report_type=report_type,
        date_from=date_from,
        date_to=date_to,
        status_filter=status_filter,
        direction_filter=direction_filter,
        author_filter=author_filter,
        selected_columns=columns,
    )

    return templates.TemplateResponse(
        request=request,
        name="reports.html",
        context={
            "user": user,
            "report_type": report_type,
            "report_presets": REPORT_PRESETS,
            "report_columns": REPORT_COLUMNS,
            "statuses": TICKET_STATUSES,
            "directions": directions,
            "filters": {
                "date_from": date_from,
                "date_to": date_to,
                "status_filter": status_filter,
                "direction_filter": direction_filter,
                "author_filter": author_filter,
                "columns": report_data["selected_columns"],
            },
            "report_data": report_data,
        },
    )


@app.get("/reports/export")
def export_report_excel(
    report_type: str = Query("summary"),
    date_from: str = Query(""),
    date_to: str = Query(""),
    status_filter: str = Query(""),
    direction_filter: str = Query(""),
    author_filter: str = Query(""),
    columns: list[str] = Query(default=[]),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    report_data = build_report_data(
        db=db,
        report_type=report_type,
        date_from=date_from,
        date_to=date_to,
        status_filter=status_filter,
        direction_filter=direction_filter,
        author_filter=author_filter,
        selected_columns=columns,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    ws["A1"] = "Отчёт по заявкам"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Всего заявок"
    ws["B3"] = report_data["total"]

    current_row = 5

    if report_type == "summary":
        ws.cell(row=current_row, column=1, value="Сводка по статусам")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        ws.append(["Статус", "Количество"])
        for status, count in report_data["status_stats"].items():
            ws.append([status, count])

        current_row = ws.max_row + 2
        ws.cell(row=current_row, column=1, value="Сводка по направлениям")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        ws.append(["Направление", "Количество"])
        for direction, count in report_data["direction_stats"].items():
            ws.append([direction, count])

        current_row = ws.max_row + 2
        ws.cell(row=current_row, column=1, value="Сводка по пользователям")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        ws.append(["Пользователь", "Количество"])
        for author, count in report_data["author_stats"].items():
            ws.append([author, count])

    elif report_type == "by_status":
        ws.append(["Статус", "Количество"])
        for status, count in report_data["status_stats"].items():
            ws.append([status, count])

    elif report_type == "by_direction":
        ws.append(["Направление", "Количество"])
        for direction, count in report_data["direction_stats"].items():
            ws.append([direction, count])

    elif report_type == "by_author":
        ws.append(["Пользователь", "Количество"])
        for author, count in report_data["author_stats"].items():
            ws.append([author, count])

    else:
        selected_columns = report_data["selected_columns"]
        headers = [REPORT_COLUMNS.get(column, column) for column in selected_columns]
        ws.append(headers)

        for row in report_data["rows"]:
            ws.append([row.get(column, "") for column in selected_columns])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top")
            if cell.row in [1, 3, 5]:
                cell.font = Font(bold=True)

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 3, 45)

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = "report.xlsx"

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        },
    )
